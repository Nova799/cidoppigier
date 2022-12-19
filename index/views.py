from datetime import date
from django.shortcuts import render, redirect
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from .forms import Login, AddUser, ModifyUser, Search, AddDoc, MdpForgot, ModifyDoc, AddEmprunt, xlsFile, Appreciation, exportUser, firstLogin, exportDoc
from .models import Emprunts, Users, Appreciations, Documents, xls
from random import choice
from hashlib import sha1
from PyPDF2 import PdfFileReader, PdfFileWriter
from openpyxl import load_workbook, Workbook
import matplotlib.pyplot as plt, numpy as np, unidecode as uni, re

areConnceted = [Users.objects.get(loginMail = "casanovakouakanou@gmail.com")]
# areConnceted = []
err_dict = {}
random_word = ["buenosaires", "azertyuiop", "qwertyuiop", "pigierbenin", "cidopigier"]

# Statics functions

def makeExportXLS(groupOfData, datas, saveDir = "cidop/index/templates/assets/xls_file/"):
    if groupOfData == "Utilisateurs":
        if datas == "state":
            states = {"Actifs": Users.objects.filter(state = True).order_by("nom"), "Inactifs": Users.objects.filter(state = False).order_by("nom")}
            wb = Workbook()
            ws = wb.active
            ws.title = "Actifs"
            cols = ["N°", "Nom et prénom(s)", "Adresse de connexion", "Adresse Mail", "Filière", "Type d'utilisateur", "Mot de passe", "Elligibilité"]
            for i in range(len(cols)):
                ws.cell(row = 1, column= i+1, value=cols[i])
            if len(states["Actifs"]) > 0:
                for i in range(len(states["Actifs"])):
                    ws.cell(row = i+2, column= 1, value=i+1)
                    ws.cell(row = i+2, column= 2, value=f'{states["Actifs"][i].nom} {states["Actifs"][i].prenom}')
                    ws.cell(row = i+2, column= 3, value=f'{states["Actifs"][i].loginMail}')
                    ws.cell(row = i+2, column= 4, value=f'{states["Actifs"][i].addrMail}')
                    ws.cell(row = i+2, column= 5, value=f'{states["Actifs"][i].filiere}')
                    ws.cell(row = i+2, column= 6, value="Utilisateur" if states["Actifs"][i].typeOfUser == "user" else "Administrateur")
                    ws.cell(row = i+2, column= 7, value=f'{states["Actifs"][i].mdp}')
                    ws.cell(row = i+2, column= 8, value="Elligible" if states["Actifs"][i].elligible else "Non elligible")
            ws = wb.create_sheet("Inactifs")
            for i in range(len(cols)):
                ws.cell(row = 1, column= i+1, value=cols[i])
            if len(states["Inactifs"]) > 0:
                for i in range(len(states["Inactifs"])):
                    ws.cell(row = i+2, column= 1, value=i+1)
                    ws.cell(row = i+2, column= 2, value=f'{states["Inactifs"][i].nom} {states["Inactifs"][i].prenom}')
                    ws.cell(row = i+2, column= 3, value=f'{states["Inactifs"][i].loginMail}')
                    ws.cell(row = i+2, column= 4, value=f'{states["Inactifs"][i].addrMail}')
                    ws.cell(row = i+2, column= 5, value=f'{states["Inactifs"][i].filiere}')
                    ws.cell(row = i+2, column= 6, value="Utilisateur" if states["Inactifs"][i].typeOfUser == "user" else "Administrateur")
                    ws.cell(row = i+2, column= 7, value=f'{states["Inactifs"][i].mdp}')
                    ws.cell(row = i+2, column= 8, value="Elligible" if states["Inactifs"][i].elligible else "Non elligible")
            wb.save(f"{saveDir}export_{groupOfData}_{datas}.xlsx")
            return True
        elif datas == "fil":
            data, fil = [], {}
            for i in Users.objects.all():
                if not i.filiere in data:
                    data.append(i.filiere)
            for i in data:
                fil[i] = Users.objects.filter(filiere = i).order_by("nom")
            wb = Workbook()
            ws = wb.active
            ws.title = list(fil.keys())[0]
            cols = ["N°", "Nom et prénom(s)", "Adresse de connexion", "Adresse Mail", "Elligibilité", "Type d'utilisateur", "Mot de passe", "État"]
            for i in range(len(cols)):
                ws.cell(row = 1, column= i+1, value=cols[i])
            if len(list(fil.keys())) > 0:
                for i in range(len(list(fil.keys())[0])):
                    ws.cell(row = i+2, column= 1, value=i+1)
                    ws.cell(row = i+2, column= 2, value=f'{fil[list(fil.keys())[0]][i].nom} {fil[list(fil.keys())[0]][i].prenom}')
                    ws.cell(row = i+2, column= 3, value=f'{fil[list(fil.keys())[0]][i].loginMail}')
                    ws.cell(row = i+2, column= 4, value=f'{fil[list(fil.keys())[0]][i].addrMail}')
                    ws.cell(row = i+2, column= 5, value=f'{"Elligible" if  fil[list(fil.keys())[0]][i].elligible else "Non elligible"}')
                    ws.cell(row = i+2, column= 6, value="Utilisateur" if fil[list(fil.keys())[0]][i].typeOfUser == "user" else "Administrateur")
                    ws.cell(row = i+2, column= 7, value=f'{fil[list(fil.keys())[0]][i].mdp}')
                    ws.cell(row = i+2, column= 8, value="Actif" if fil[list(fil.keys())[0]][i].state else "Inactif")
            if len(list(fil.keys())) > 1:
                for key in list(fil.keys())[1:]:
                    ws = wb.create_sheet(key)
                    for i in range(len(cols)):
                        ws.cell(row = 1, column= i + 1, value=cols[i])
                    for i in range(len(list(fil[key]))):
                        ws.cell(row = i+2, column= 1, value=i+1)
                        ws.cell(row = i+2, column= 2, value=f'{fil[key][i].nom} {fil[key][i].prenom}')
                        ws.cell(row = i+2, column= 3, value=f'{fil[key][i].loginMail}')
                        ws.cell(row = i+2, column= 4, value=f'{fil[key][i].addrMail}')
                        ws.cell(row = i+2, column= 5, value=f'{"Elligible" if  fil[key][i].elligible else "Non elligible"}')
                        ws.cell(row = i+2, column= 6, value="Utilisateur" if fil[key][i].typeOfUser == "user" else "Administrateur")
                        ws.cell(row = i+2, column= 7, value=f'{fil[key][i].mdp}')
                        ws.cell(row = i+2, column= 8, value="Actif" if fil[key][i].state else "Inactif")
            wb.save(f"{saveDir}export_{groupOfData}_{datas}.xlsx")
            return True
        elif datas == "elig":
            elig = {"Elligible": Users.objects.filter(elligible = True).order_by("nom"), "Non Elligible": Users.objects.filter(elligible = False).order_by("nom")}
            wb = Workbook()
            ws = wb.active
            ws.title = "Elligibles"
            cols = ["N°", "Nom et prénom(s)", "Adresse de connexion", "Adresse Mail", "Filière", "Type d'utilisateur", "Mot de passe", "État"]
            for i in range(len(cols)):
                ws.cell(row = 1, column= i+1, value=cols[i])
            if len(elig["Elligible"]) > 0:
                for i in range(len(elig["Elligible"])):
                    ws.cell(row = i+2, column= 1, value=i+1)
                    ws.cell(row = i+2, column= 2, value=f'{elig["Elligible"][i].nom} {elig["Elligible"][i].prenom}')
                    ws.cell(row = i+2, column= 3, value=f'{elig["Elligible"][i].loginMail}')
                    ws.cell(row = i+2, column= 4, value=f'{elig["Elligible"][i].addrMail}')
                    ws.cell(row = i+2, column= 5, value=f'{elig["Elligible"][i].filiere}')
                    ws.cell(row = i+2, column= 6, value="Utilisateur" if elig["Elligible"][i].typeOfUser == "user" else "Administrateur")
                    ws.cell(row = i+2, column= 7, value=f'{elig["Elligible"][i].mdp}')
                    ws.cell(row = i+2, column= 8, value="Actif" if elig["Elligible"][i].state else "Inactif")
            ws = wb.create_sheet("Non elligibles")
            for i in range(len(cols)):
                ws.cell(row = 1, column= i+1, value=cols[i])
            if len(elig["Non Elligible"]) > 0:
                for i in range(len(elig["Non Elligible"])):
                    ws.cell(row = i+2, column= 1, value=i+1)
                    ws.cell(row = i+2, column= 2, value=f'{elig["Non Elligible"][i].nom} {elig["Non Elligible"][i].prenom}')
                    ws.cell(row = i+2, column= 3, value=f'{elig["Non Elligible"][i].loginMail}')
                    ws.cell(row = i+2, column= 4, value=f'{elig["Non Elligible"][i].addrMail}')
                    ws.cell(row = i+2, column= 5, value=f'{elig["Non Elligible"][i].filiere}')
                    ws.cell(row = i+2, column= 6, value="Utilisateur" if elig["Non Elligible"][i].typeOfUser == "user" else "Administrateur")
                    ws.cell(row = i+2, column= 7, value=f'{elig["Non Elligible"][i].mdp}')
                    ws.cell(row = i+2, column= 8, value="Actif" if elig["Non Elligible"][i].state else "Inactif")
            wb.save(f"{saveDir}export_{groupOfData}_{datas}.xlsx")
            return True
        elif datas == "typ":
            typ = {"Utilisateurs": Users.objects.filter(typeOfUser = "user").order_by("nom"), "Administrateurs": Users.objects.filter(typeOfUser = "admin").order_by("nom")}
            wb = Workbook()
            ws = wb.active
            ws.title = "Utilisateurs"
            cols = ["N°", "Nom et prénom(s)", "Adresse de connexion", "Adresse Mail", "Filière", "Elligibilité", "Mot de passe", "État"]
            for i in range(len(cols)):
                ws.cell(row = 1, column= i+1, value=cols[i])
            if len(typ["Utilisateurs"]) > 0:
                for i in range(len(typ["Utilisateurs"])):
                    ws.cell(row = i+2, column= 1, value=i+1)
                    ws.cell(row = i+2, column= 2, value=f'{typ["Utilisateurs"][i].nom} {typ["Utilisateurs"][i].prenom}')
                    ws.cell(row = i+2, column= 3, value=f'{typ["Utilisateurs"][i].loginMail}')
                    ws.cell(row = i+2, column= 4, value=f'{typ["Utilisateurs"][i].addrMail}')
                    ws.cell(row = i+2, column= 5, value=f'{typ["Utilisateurs"][i].filiere}')
                    ws.cell(row = i+2, column= 6, value="Elligible" if typ["Utilisateurs"][i].elligible else "Non elligible")
                    ws.cell(row = i+2, column= 7, value=f'{typ["Utilisateurs"][i].mdp}')
                    ws.cell(row = i+2, column= 8, value="Actif" if typ["Utilisateurs"][i].state else "Inactif")
            ws = wb.create_sheet("Administrateurs")
            for i in range(len(cols)):
                ws.cell(row = 1, column= i+1, value=cols[i])
            if len(typ["Administrateurs"]) > 0:
                for i in range(len(typ["Administrateurs"])):
                    ws.cell(row = i+2, column= 1, value=i+1)
                    ws.cell(row = i+2, column= 2, value=f'{typ["Administrateurs"][i].nom} {typ["Administrateurs"][i].prenom}')
                    ws.cell(row = i+2, column= 3, value=f'{typ["Administrateurs"][i].loginMail}')
                    ws.cell(row = i+2, column= 4, value=f'{typ["Administrateurs"][i].addrMail}')
                    ws.cell(row = i+2, column= 5, value=f'{typ["Administrateurs"][i].filiere}')
                    ws.cell(row = i+2, column= 6, value="Elligible" if typ["Administrateurs"][i].elligible else "Non elligible")
                    ws.cell(row = i+2, column= 7, value=f'{typ["Administrateurs"][i].mdp}')
                    ws.cell(row = i+2, column= 8, value="Actif" if typ["Administrateurs"][i].state else "Inactif")
            wb.save(f"{saveDir}export_{groupOfData}_{datas}.xlsx")
            return True
        elif datas == "all":
            wb = Workbook()
            ws = wb.active
            ws.title = "Utilisateurs"
            data = Users.objects.order_by("nom")
            cols = ["N°", "Nom et prénom(s)", "Adresse de connexion", "Adresse Mail", "Filière", "Mot de passe", "Elligibilité", "État", "Type d'utilisateurs"]
            if len(data) > 0:
                for i in range(len(data)):
                    ws.cell(row = i+2, column= 1, value=i+1)
                    ws.cell(row = i+2, column= 2, value=f'{data[i].nom} {data[i].prenom}')
                    ws.cell(row = i+2, column= 3, value=f'{data[i].loginMail}')
                    ws.cell(row = i+2, column= 4, value=f'{data[i].addrMail}')
                    ws.cell(row = i+2, column= 5, value=f'{data[i].filiere}')
                    ws.cell(row = i+2, column= 6, value=f'{data[i].mdp}')
                    ws.cell(row = i+2, column= 7, value="Elligible" if data[i].elligible else "Non elligible")
                    ws.cell(row = i+2, column= 8, value="Actif" if data[i].state else "Inactif")
                    ws.cell(row = i+2, column= 9, value="Elligible" if data[i].elligible else "Non elligible")
            wb.save(f"{saveDir}export_{groupOfData}_{datas}.xlsx")
            return True
        else:
            return False
    elif groupOfData == "Documents":
        if datas == "state":
            states = {"Actifs": Documents.objects.filter(existe = True).order_by("titre"), "Inactifs": Documents.objects.filter(existe = False).order_by("titre")}
            wb = Workbook()
            ws = wb.active
            ws.title = "Actifs"
            cols = ["N°", "Titre", "Catégorie", "Résumé", "Nom de l'auteur", "Type", "Emplacement", "Disponibilité"]
            for i in range(len(cols)):
                ws.cell(row = 1, column= i+1, value=cols[i])
            if len(states["Actifs"]) > 0:
                for i in range(len(states["Actifs"])):
                    ws.cell(row = i+2, column= 1, value=i+1)
                    ws.cell(row = i+2, column= 2, value=f'{states["Actifs"][i].titre}')
                    ws.cell(row = i+2, column= 3, value=f'{states["Actifs"][i].typeOfDoc}')
                    ws.cell(row = i+2, column= 4, value=f'{states["Actifs"][i].resume}')
                    ws.cell(row = i+2, column= 5, value=f'{states["Actifs"][i].auteur}')
                    ws.cell(row = i+2, column= 6, value="Physique" if states["Actifs"][i].physique else "Numérique")
                    ws.cell(row = i+2, column= 7, value=f'Rangée {states["Actifs"][i].emplacement}' if states["Actifs"][i].physique else "-")
                    ws.cell(row = i+2, column= 8, value=f'Disponible' if states["Actifs"][i].disponible else "Non disponible")
            ws = wb.create_sheet("Inactifs")
            for i in range(len(cols)):
                ws.cell(row = 1, column= i+1, value=cols[i])
            if len(states["Inactifs"]) > 0:
                for i in range(len(states["Inactifs"])):
                    ws.cell(row = i+2, column= 1, value=i+1)
                    ws.cell(row = i+2, column= 2, value=f'{states["Inactifs"][i].titre}')
                    ws.cell(row = i+2, column= 3, value=f'{states["Inactifs"][i].typeOfDoc}')
                    ws.cell(row = i+2, column= 4, value=f'{states["Inactifs"][i].resume}')
                    ws.cell(row = i+2, column= 5, value=f'{states["Inactifs"][i].auteur}')
                    ws.cell(row = i+2, column= 6, value="Physique" if states["Inactifs"][i].physique else "Numérique")
                    ws.cell(row = i+2, column= 7, value=f'Rangée {states["Inactifs"][i].emplacement}' if states["Inactifs"][i].physique else "-")
                    ws.cell(row = i+2, column= 8, value=f'Disponible' if states["Inactifs"][i].disponible else "Non disponible")
            wb.save(f"{saveDir}export_{groupOfData}_{datas}.xlsx")
            return True
        elif datas == "categ":pass
        elif datas == "typ":
            states = {"Physique": Documents.objects.filter(physique = True).order_by("titre"), "Numérique": Documents.objects.filter(physique = False).order_by("titre")}
            wb = Workbook()
            ws = wb.active
            ws.title = "Physiques"
            cols = ["N°", "Titre", "Catégorie", "Résumé", "Nom de l'auteur", "État", "Emplacement", "Disponibilité"]
            for i in range(len(cols)):
                ws.cell(row = 1, column= i+1, value=cols[i])
            if len(states["Physique"]) > 0:
                for i in range(len(states["Physique"])):
                    ws.cell(row = i+2, column= 1, value=i+1)
                    ws.cell(row = i+2, column= 2, value=f'{states["Physique"][i].titre}')
                    ws.cell(row = i+2, column= 3, value=f'{states["Physique"][i].typeOfDoc}')
                    ws.cell(row = i+2, column= 4, value=f'{states["Physique"][i].resume}')
                    ws.cell(row = i+2, column= 5, value=f'{states["Physique"][i].auteur}')
                    ws.cell(row = i+2, column= 6, value="Actif" if states["Physique"][i].existe else "Inactif")
                    ws.cell(row = i+2, column= 7, value=f'Rangée {states["Physique"][i].emplacement}' if states["Physique"][i].physique else "-")
                    ws.cell(row = i+2, column= 8, value=f'Disponible' if states["Physique"][i].disponible else "Non disponible")
            ws = wb.create_sheet("Numériques")
            for i in range(len(cols)):
                ws.cell(row = 1, column= i+1, value=cols[i])
            if len(states["Numérique"]) > 0:
                for i in range(len(states["Numérique"])):
                    ws.cell(row = i+2, column= 1, value=i+1)
                    ws.cell(row = i+2, column= 2, value=f'{states["Numérique"][i].titre}')
                    ws.cell(row = i+2, column= 3, value=f'{states["Numérique"][i].typeOfDoc}')
                    ws.cell(row = i+2, column= 4, value=f'{states["Numérique"][i].resume}')
                    ws.cell(row = i+2, column= 5, value=f'{states["Numérique"][i].auteur}')
                    ws.cell(row = i+2, column= 6, value="Actif" if states["Numérique"][i].existe else "Inactif")
                    ws.cell(row = i+2, column= 7, value=f'Rangée {states["Numérique"][i].emplacement}' if states["Numérique"][i].physique else "-")
                    ws.cell(row = i+2, column= 8, value=f'Disponible' if states["Numérique"][i].disponible else "Non disponible")
            wb.save(f"{saveDir}export_{groupOfData}_{datas}.xlsx")
            return True
        elif datas == "disp":
            pass
        elif datas == "auteur":pass
        elif datas == "avg":pass
        elif datas == "all":
            pass
        else:
            return False
    elif groupOfData == "Emprunts":
        return False
    else:
        return False
    return False

def reserved(user_idf, doc_idfy):
    try:
        user = userAllInfoFromUserIdf(user_idf)
    except:return False
    else:
        try:
            doc = Documents.objects.get(idfy=doc_idfy)
        except:return False
        else:
            if user.idfy in doc.reservedby:return True
            else: return False
        
def paginate(obj_list, page_number = 0):
    paginator = Paginator(obj_list, 15)
    page_number = page_number
    page_obj = paginator.get_page(page_number)
    return page_obj

def autoAdd(fle_path, import_group):
    wb = load_workbook(filename=fle_path)
    ws = wb.active
    max_row = ws.max_row
    max_column = ws.max_column
    if import_group == "Utilisateurs":
        if ws.cell(row=1, column=2).value.lower().__contains__("nom"):
            for row in ws.iter_rows(min_row=2, min_col=2, max_row=max_row):
                msg, tom = createUser(nom=row[0].value, prenom=row[1].value, addrMail=row[2].value, filiere=row[3].value)
            return ("tip", "Les utilisateurs ont été ajouté avec succès")
        else:
            return ("info", "Les données du fichier sont invalides")
    elif import_group == "Documents":
        if ws.cell(row=1, column=2).value.lower().__contains__("titre"):
            for row in ws.iter_rows(min_row=2, min_col=2, max_row=max_row):
                doc = Documents()
                doc.titre = row[0].value
                doc.resume = row[1].value
                doc.typeOfDoc = row[2].value
                doc.auteur = row[3].value
                doc.emplacement = row[4].value
                doc.isbn = row[5].value
                doc.physique = True
                global random_word
                idfy = "{}{}{}".format(doc.titre, choice(range(100)), choice(random_word), doc.emplacement)
                idfy = idfy.encode()
                idfy = sha1(idfy)
                doc.idfy = idfy.hexdigest()
                doc.save()
                
            return ("tip", "Les documents ont été ajouté avec succès")
        else:
            return ("info", "Les données du fichier sont invalides")
    else:return ("info", "Une erreur inattendue est survenue. Veuillez recharger la page")

def isConnected(user_idf):
    global areConnceted
    if userAllInfoFromUserIdf(user_idf) in areConnceted:return True
    else:return False

def isAdmin(user_idf):
    user = userAllInfoFromUserIdf(user_idf)
    if not user == None:
        if user.typeOfUser == "admin":return True
        else:return False
    else:return False

def idfyFromIdf(user_idf=str):
    return user_idf.split("k")[1]

def verifyUserExist(user_idfy=str):
    try:Users.objects.get(idfy=user_idfy)
    except:return False
    else:return True

def userAllInfoFromUserIdf(user_idf=str):
    user_idfy = user_idf.split("k")[1]
    if verifyUserExist(user_idfy=user_idfy):return Users.objects.get(idfy=user_idfy)
    else:return None

def makeUserIdf(user_idfy=str):
    if verifyUserExist(user_idfy=user_idfy):
        supl = "cid0p1@".encode()
        supl = sha1(supl).hexdigest()
        return f"{supl}k{user_idfy}"
    else:return None

def makeloginMail(nom = str, prenom = str):
    nom = uni.unidecode(nom.strip())
    prenom = uni.unidecode(prenom.strip())
    nom = nom.replace("-", "")
    prenom = prenom.replace("-", "")
    nom = nom.replace(".", "")
    prenom = prenom.replace(".", "")
    prenom = prenom.split(" ")
    loginMail = f"{nom.lower()}.{prenom[0].lower()}@pigierbenin.com"
    try:
        Users.objects.get(loginMail = loginMail)
    except Users.DoesNotExist:
        return f"{nom.lower()}.{prenom[0].lower()}@pigierbenin.com"
    else:
        if len(prenom) > 1:return f"{nom.lower()}.{prenom[0].lower()}.{prenom[1].lower()}@pigierbenin.com"
        else:return f"{nom.lower()}.{prenom[0].lower()}@pigierbenin.com"

def createUser(nom, prenom, addrMail, typeOfUser="user", filiere="-"):
    if addrMail.strip() == "":
        return ("Veuillez remplir tout les champs", "info")
    loginMail = makeloginMail(nom, prenom)
    try:
        user = Users.objects.get(state=True, nom=nom, prenom=prenom, filiere=filiere, loginMail=loginMail)
    except:
        global random_word
        mdp = "{}{}{}".format(loginMail, choice(range(100)), choice(random_word))
        mdp = mdp.encode()
        mdp = sha1(mdp)
        if typeOfUser == "admin":
            user = Users(idfy=mdp.hexdigest(), nom=nom, prenom=prenom, loginMail=loginMail, addrMail=addrMail, typeOfUser="admin", mdp=mdp.hexdigest()[:8], filiere="-")
        else:
            user = Users(idfy=mdp.hexdigest(), nom=nom, prenom=prenom, loginMail=loginMail, addrMail=addrMail, typeOfUser="user", mdp=mdp.hexdigest()[:8], filiere=filiere)
        user.save()
        del mdp, user
        return ("L'utilisateur a été ajouté avec succès", "tip")
    else:
        return ("L'utilisateur existe déjà", "info")

def encryptPdf(pdfPath = str, idfy = str):
    result = PdfFileWriter()
    file = PdfFileReader(pdfPath)
    length = file.numPages
    for i in range(length):
        pages = file.getPage(i)
        result.addPage(pages)
    password = idfy[3:16]
    result.encrypt(password)
    with open(pdfPath,'wb') as f:
        result.write(f)

def dateToTuple(date = date):
    return (date.day, date.month, date.year)

def returnMaxDay(month = int, year = date.today().year):
    month, year = int(month), int(year)
    vals = {
        "31": [1, 3, 5, 7, 8, 10, 12],
        "30": [4, 6, 9, 11],
        }
    for key in vals.keys():
        if month == 2:
            if year > 0:
                if year % 4 == 0:return 29
                else:return 28
            else:
                return 28
        elif month in vals[key]:
            return int(key)
    return 0

def dateToStr(date = date):
    return f"{date.day}/{date.month}/{date.year}"

def add2Weeks(today = date):
    d, m, y = dateToTuple(today)
    d, m, y = int(d), int(m), int(y)
    maxDay = returnMaxDay(m, y)
    if d + 14 > maxDay:
        if m < 12:
            m = m + 1
            d = (d + 14) - maxDay
            return date(y, m, d)
        elif m == 12:
            y = y + 1
            m = 1
            d = (d + 14) - maxDay
            return date(y, m, d)
    else:
        return date(y, m, d + 14)

def retard(dateEmprunt = date, dateLimit = date):
    dateLimit = add2Weeks(dateEmprunt)
    today = date.today()
    if dateLimit.year > today.year:
        return 0
    elif dateLimit.year == today.year:
        if dateLimit.month > today.month:
            return 0
        elif dateLimit.month == today.month:
            if dateLimit.day >= today.day:
                return 0
            else:
                return today.day - dateLimit.day
        else:
            returnVal = 0
            m = dateLimit.month
            while m < today.month:
                if returnVal == 0:
                    returnVal += returnMaxDay(m, dateLimit.year) - dateLimit.day
                else:
                    returnVal += returnMaxDay(m, dateLimit.year)
                m += 1
            return returnVal + today.day
    elif dateLimit.year < today.year:
        returnVal = 0
        m, y = dateLimit.month, dateLimit.year
        returnVal += returnMaxDay(m, y) - dateLimit.day       
        while y < today.year:
            if m == 12:
                m = 1
                y += 1
            elif m < 12:
                m += 1
                returnVal += returnMaxDay(m)
        if m == today.month:
            return returnVal + today.day
        else:
            while m < today.month:
                returnVal += returnMaxDay(m, y)
                m += 1
            return returnVal + today.day
    return 0

def calculatePenality (retard = int):
    return retard * 250

# Create your views here.

# Login / Logout

def login(request):
    projet_name = 'CIDOP'
    global err_dict
    global areConnceted
    if request.method == "POST":
        form = Login(request.POST)
        if form.is_valid():
            loginMail = form.cleaned_data['loginMail']
            mdp = form.cleaned_data['mdp']
            if not loginMail in err_dict.keys():
                err_dict[loginMail] = 0
            try:
                user = Users.objects.get(loginMail=loginMail, mdp=mdp, state=True)
            except:
                msg = "L'utilisateur est introuvable"
                err_dict[loginMail] += 1
                if not err_dict[loginMail] == 3:
                    return render(request, "login.html", locals())
                err_dict[loginMail] = 0
                return render(request, "user_3_errors.html", locals())
            else:
                err_dict.pop(loginMail)
                areConnceted.append(user)
                if user.mdp == user.idfy[3:11]:
                    return redirect(f"/firstLogin/{makeUserIdf(user.idfy)}")
                if user.typeOfUser != "admin":
                    return redirect(f"/Accueil/{makeUserIdf(user.idfy)}")
                else:
                    return redirect(f"/Admin/{makeUserIdf(user.idfy)}/Utilisateurs/")
    else:
        form = Login()
    return render(request, "login.html", locals())

def firstConn(request, user_idf):
    projet_name = "CIDOP"
    if isConnected(user_idf):
        user = userAllInfoFromUserIdf(user_idf)
        if user == None:return redirect("/")
        else:
            if user.mdp == user.idfy[3:11]:
                if request.method == "POST":
                    form = firstLogin(request.POST)
                    if form.is_valid():
                        mdp1 = form.cleaned_data['mdp1']
                        mdp2 = form.cleaned_data['mdp2']
                        if mdp1 != mdp2:msg = "Les mots de passes ne concordent pas"
                        else:
                            if len(mdp1) < 8:msg = "Mots de passes trop court."
                            else:
                                if re.match(".+[A-Za-z]+.+", mdp1) == None:msg = "Doit contenir au moins une lettre"
                                else:
                                    if re.match(".+[0-9]+.+", mdp1) == None:msg = "Doit contenir au moins un chiffre"
                                    else:
                                        if re.match(".+[\@\_\+\.\-\#]+.+", mdp1) == None:msg = "Doit contenir au moins '@', '_', '+', '.', '#', ou '-'"
                                        else:
                                            user.mdp = mdp1
                                            user.save()
                                            if user.typeOfUser != "admin":return redirect(f"/Accueil/{user_idf}")
                                            else:return redirect(f"/Admin/{user_idf}/Utilisateurs/")
                else:form = firstLogin()
            else:return redirect(f"/")
        return render(request, "firstLogin.html", locals())
    else:return redirect(f"/")

def logout(request, user_idf):
    global areConnceted
    if isConnected(user_idf):
        areConnceted.remove(userAllInfoFromUserIdf(user_idf))
        return redirect("/")

def mdpForgot(request):
    projet_name = "CIDOP"
    if request.method == "POST":
        form = MdpForgot(request.POST)
        if form.is_valid():
            loginMail = form.cleaned_data["loginMail"]
            recover = form.cleaned_data["recover"]
            options = form.cleaned_data["options"]
            try:user = Users.objects.get(loginMail = loginMail.strip())
            except:typeOfMsg, msg = ("txt-a", "L'identifiant est incorrect")
            else:
                if user.idfy[3:11] == recover:    
                    user.mdp = user.idfy[3:11]
                    user.save()
                    if not options == None:
                        if options == 'all':
                            global areConnceted
                            try:
                                areConnceted.remove(user)
                            except:
                                pass
                        tom, msg = ("txt-s", "Le mot de passe a été renouvelé. \nConnectez-vous avec votre clef de récupération")
                    else:
                        tom, msg = ("txt-a", "Aucune option n'a été choisie")
    else:
        form = MdpForgot()
    return render(request, 'mdpForgot.html', locals())

def home(request, user_idf):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        adm = False
        if isAdmin(user_idf):
            adm = True
        books_target = ""
        for i in Documents.objects.filter(existe = True):
            books_target += f"{i.titre} - {i.auteur}##"
        if request.GET.get("search") or not request.GET.get("search") in [None, ""]:
            doc = Documents.objects.get(titre = request.GET.get("search").split("-")[0].strip(), auteur = request.GET.get("search").split("-")[1].strip())
            return redirect(f"/DocInfo/{user_idf}/{doc.idfy}/")
        return render(request, 'index.html', locals())
    else:
        return redirect("/")

def preferences(request, user_idf):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        adm = False
        if isAdmin(user_idf):
            adm = True
        mine = []
        if len(Appreciations.objects.all()) > 0:
            for i in Appreciations.objects.all():
                if user_idf == makeUserIdf(i.user.idfy):
                    if i.note >= 12:
                        if not i.book.typeOfDoc in mine:
                            mine.append(i.book.typeOfDoc)
        return render(request, 'preferences.html', locals())
    else:
        return redirect("/")

def recommandations(request, user_idf):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        adm = False
        if isAdmin(user_idf):
            adm = True
        mine, docs = [], []
        if len(Appreciations.objects.all()) > 0:
            for i in Appreciations.objects.filter(user = userAllInfoFromUserIdf(user_idf)):
                    if i.note >= 12:
                        if not i.book.typeOfDoc in mine:
                            mine.append(i.book.typeOfDoc)
            for i in mine:
                for x in Documents.objects.filter(typeOfDoc = i):
                    docs.append(x)
            page_obj = paginate(docs, request.GET.get("page"))
        return render(request, 'recommandations.html', locals())
    else:
        return redirect("/")

def about(request, user_idf):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        books_target = ""
        adm = False
        if isAdmin(user_idf):
            adm = True
        for i in Documents.objects.filter(existe = True):
            books_target += f"{i.titre} - {i.auteur}##"
        if request.GET.get("search") or not request.GET.get("search") in [None, ""]:
            doc = Documents.objects.get(titre = request.GET.get("search").split("-")[0].strip(), auteur = request.GET.get("search").split("-")[1].strip())
            return redirect(f"/DocInfo/{user_idf}/{doc.idfy}/")
        return render(request, 'about.html', locals())
    else:
        return redirect("/")

def all_docs(request, user_idf):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        books_target = ""
        adm = False
        if isAdmin(user_idf):
            adm = True
        for i in Documents.objects.filter(existe = True):
            books_target += f"{i.titre} - {i.auteur}##"
        if request.GET.get("search") or not request.GET.get("search") in [None, ""]:
            doc = Documents.objects.get(titre = request.GET.get("search").split("-")[0].strip(), auteur = request.GET.get("search").split("-")[1].strip())
            return redirect(f"/DocInfo/{user_idf}/{doc.idfy}/")
        docs = Documents.objects.all()
        page_obj = paginate(docs, request.GET.get("page"))
        return render(request, 'all_docs.html', locals())
    else:
        return redirect("/")

def most_noted(request, user_idf):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        books_target = ""
        adm = False
        if isAdmin(user_idf):
            adm = True
        for i in Documents.objects.filter(existe = True):
            books_target += f"{i.titre} - {i.auteur}##"
        if request.GET.get("search") or not request.GET.get("search") in [None, ""]:
            doc = Documents.objects.get(titre = request.GET.get("search").split("-")[0].strip(), auteur = request.GET.get("search").split("-")[1].strip())
            return redirect(f"/DocInfo/{user_idf}/{doc.idfy}/")
        docs = []
        for i in Documents.objects.filter(physique = False, existe = True):
            if i .avg >= 12:
                if not i in docs:
                    docs.append(i)
        page_obj = paginate(docs, request.GET.get("page"))
        return render(request, 'most_noted.html', locals())
    else:
        return redirect("/")

def recents(request, user_idf):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        adm = False
        if isAdmin(user_idf):
            adm = True
        books_target = ""
        for i in Documents.objects.filter(existe = True):
            books_target += f"{i.titre} - {i.auteur}##"
        if request.GET.get("search") or not request.GET.get("search") in [None, ""]:
            doc = Documents.objects.get(titre = request.GET.get("search").split("-")[0].strip(), auteur = request.GET.get("search").split("-")[1].strip())
            return redirect(f"/DocInfo/{user_idf}/{doc.idfy}/")
        docs = Documents.objects.order_by("-id")
        docs = docs[:15]
        if request.GET.get("info") or not request.GET.get("info") in [None, ""]:
            doc_my = Documents.objects.get(idfy = request.GET.get("info"))
            passwd = doc_my.idfy[3:16]
        if request.GET.get("note") or not request.GET.get("note") in [None, ""]:
            Appreciations(user=userAllInfoFromUserIdf(user_idf), book=doc_my, note=request.GET.get("note")).save()
        return render(request, 'recents.html', locals())
    else:
        return redirect("/")

def adm_users(request, user_idf):
    projet_name = 'CIDOP'
    global areConnected
    message, typeOfMsg = ("", "")
    if isConnected(user_idf):
        if isAdmin(user_idf):
            filieres = ""
            for i in Users.objects.all():
                if not i.filiere in filieres:
                    filieres += i.filiere + "##"
            if request.method == "POST":
                form = AddUser(request.POST)
                if form.is_valid():
                    nom = form.cleaned_data['nom']
                    prenom = form.cleaned_data['prenom']
                    addrMail = form.cleaned_data['addrMail']
                    typeOfUser = form.cleaned_data['typeOfUser']
                    filiere = form.cleaned_data['filiere']
                    msg, tom = createUser(nom=nom, prenom=prenom, addrMail=addrMail, typeOfUser=typeOfUser, filiere=filiere)
            else:
                form = AddUser()
        else:
            return redirect(f"/Accueil/{user_idf}/")
    else:
        return redirect("/")
    return render(request, 'admins/index.html', locals())

def adm_users_modify(request, user_idf, idfy_user):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        if request.method == "POST":
            form = ModifyUser(request.POST)
            if form.is_valid():
                nom = form.cleaned_data['nom']
                prenom = form.cleaned_data['prenom']
                loginMail = form.cleaned_data['loginMail']

                user = Users.objects.get(idfy=idfy_user, state=True)
                user.nom = nom
                user.prenom = prenom
                user.loginMail = user.loginMail
                try:
                    user.save()
                except:
                    message = "Cette adresse est déjà utilisée"
                    typeOfMsg = "txt-a"
                else:
                    message = "La modification a été effectuée avec succès."
                    typeOfMsg = "txt-s"
        else:
            form = ModifyUser()
        return render(request, 'data_mod.html', locals())
    else:
        return redirect("/")

def adm_users_del(request, user_idf, idfy_user):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        try:
            user = Users.objects.get(idfy=idfy_user, state=True)
        except:
            pass
        else:
            user.state = False
            user.save()
            return redirect("../../")
        return render(request, 'user_del.html', locals())
    else:
        return redirect("/")

def adm_docs(request, user_idf):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        if isAdmin(user_idf):
            auteurs = ""
            for i in Documents.objects.all():
                if not i.auteur in auteurs:
                    auteurs += f"{i.auteur}##"
            if request.method == "POST":
                form = AddDoc(request.POST, request.FILES)
                if form.is_valid():
                    titre = form.cleaned_data["titre"]
                    typeOfDoc = form.cleaned_data["typeOfDoc"]
                    resume = form.cleaned_data["resume"]
                    auteur = form.cleaned_data["auteur"]
                    emplacement = form.cleaned_data["emplacement"]
                    physique = form.cleaned_data["physique"]
                    filePath = form.cleaned_data["filePath"]
                    isbn = form.cleaned_data['isbn']
                    doc = Documents()
                    global random_word
                    idfy = "{}{}{}".format(titre, choice(range(100)), choice(random_word), emplacement)
                    idfy = idfy.encode()
                    idfy = sha1(idfy)
                    doc.idfy = idfy.hexdigest()
                    doc.titre = titre
                    doc.auteur = auteur
                    doc.typeOfDoc = str(typeOfDoc).lower()
                    doc.resume = resume
                    if isbn.strip() != "":doc.isbn = isbn
                    else:doc.isbn = "-"
                    if physique:
                        doc.physique = True
                        doc.emplacement = emplacement
                    else:
                        doc.physique = False
                        doc.file_path = filePath
                        if not doc.file_path.path.split(".")[-1] == "pdf":
                            tom = "info"
                            msg = "Le format du fichier est invalide (.pdf requis)"
                            return render(request, 'admins/docs_index.html', locals())
                        doc.disponible = True
                    try:
                        doc.save()
                        if not physique:
                            encryptPdf(doc.file_path.path, idfy.hexdigest())
                    except Exception as e:
                        tom = "info"
                        msg = "Une erreur inattendue est survenue"
                    else:
                        tom = "tip"
                        msg = "Le document a été ajouté avec succès"
            else:
                form = AddDoc()
            return render(request, 'admins/docs_index.html', locals())
        else:
            return redirect(f"/Accueil/{user_idf}")
    else:
        return redirect("/")

def search(request, user_idf, searchGroup):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        user, doc, emp = Users(), Documents(), Emprunts()
        user.idfy, doc.idfy, emp.idfy = "0", "0", "0"
        targets = ""
        if searchGroup in ["Utilisateurs"]:
            for i in Users.objects.filter(state = True):
                targets += f"{i.loginMail}##"
        elif searchGroup == "Documents":
            for i in Documents.objects.filter(existe = True):
                targets += f"{i.titre} - {i.auteur}##"
        elif searchGroup == "Emprunts":
            for i in Emprunts.objects.all():
                targets += f"{i.user.loginMail} - {i.book.titre}##"
        if request.method == 'GET':
            form = Search(request.GET)
            if form.is_valid():
                target = request.GET.get("target")

                if searchGroup == "Utilisateurs":
                    try:
                        user = Users.objects.get(loginMail=target)
                    except:
                        msg = "L'utilisateur n'existe pas."
                    else:
                        return render(request, "admins/search.html", locals())
                elif searchGroup == "Documents":
                    try:
                        doc = Documents.objects.get(titre=target.split("-")[0].strip(), auteur=target.split("-")[1].strip())
                    except:
                        msg = "Le document n'existe pas."
                    else:
                        return render(request, "admins/search.html", locals())
                elif searchGroup == "Emprunts":
                    try:
                        emp = Emprunts.objects.get(user = Users.objects.get(loginMail = target.split("-")[0].strip(), state = True), book = Documents.objects.get(titre = target.split("-")[1].strip()))
                    except:
                        msg = "Aucun emprunt n'est enregistré pour cet utilisateur"
                    else:
                        dteLimite = add2Weeks(emp.dateEmprunt)
                        penDay = retard(emp.dateEmprunt, dteLimite)
                        user = emp.user
                        doc = emp.book
                        return render(request, "admins/search.html", locals())
        else:
            form = Search()
        return render(request, 'admins/search.html', locals())
    else:
        return redirect("/")

def delete_target(request, user_idf, target, searchGroup):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        if searchGroup == "Utilisateurs":
            user = Users.objects.get(idfy = target, state = True)
            user.state = False
            user.save()
            return redirect(f"/Admin/{user_idf}/Utilisateurs/")
        if searchGroup == "Documents":
            doc = Documents.objects.get(idfy = target, existe = True)
            doc.existe = False
            doc.save()
            return redirect(f"/Admin/{user_idf}/Documents/")
    else:return redirect("/")

def modify_target(request, user_idf, target, searchGroup):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        if searchGroup == "Utilisateurs":
            user = Users.objects.get(idfy = target, state = True)
            if request.method == "POST":
                form = ModifyUser(request.POST)
                if form.is_valid():
                    nom = form.cleaned_data["nom"]
                    prenom = form.cleaned_data["prenom"]
                    filiere = form.cleaned_data["filiere"]
                    addrMail = form.cleaned_data["addrMail"]
                    typeOfUser = form.cleaned_data["typeOfUser"]
                    user.nom = nom
                    user.prenom = prenom
                    user.filiere = filiere
                    user.addrMail = addrMail
                    if typeOfUser:
                        user.typeOfUser = "admin"
                    else:
                        user.typeOfUser = "user"
                    user.save()
                    msg = "La modification a été un succès"
            else:
                form = ModifyUser()
            return render(request, "admins/data_mod.html", locals())
        if searchGroup == "Documents":
            doc = Documents.objects.get(idfy = target, existe = True)
            if request.method == "POST":
                form = ModifyDoc(request.POST, request.FILES)
                if form.is_valid():
                    titre = form.cleaned_data["titre"]
                    auteur = form.cleaned_data["auteur"]
                    typeOfDoc = form.cleaned_data["typeOfDoc"]
                    physique = form.cleaned_data["physique"]
                    resume = form.cleaned_data["resume"]
                    emplacement = form.cleaned_data["emplacement"]
                    filePath = form.cleaned_data["filePath"]
                    doc.titre = titre
                    doc.auteur = auteur
                    doc.typeOfDoc = typeOfDoc
                    doc.resume = resume
                if physique:
                    doc.file_path = ""
                    doc.physique = True
                    doc.emplacement = emplacement
                else:
                    doc.emplacement = ""
                    doc.physique = False
                    doc.file_path = filePath
                    doc.disponible = True
                doc.save()
                msg = "La modification a été un succès"
            else:
                form = ModifyDoc()
            return render(request, "admins/data_mod.html", locals())
    else:return redirect("/")

def adm_emp(request, user_idf):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        if isAdmin(user_idf):
            users, books = "", ""
            for i in Users.objects.filter(state = True, elligible = True):
                users += f"{i.loginMail}##"
            for i in Documents.objects.filter(existe = True, physique = True, disponible = True):
                books += f"{i.titre} - {i.auteur}##"
            today = date.today()
            dteLimite = add2Weeks(today)
            if request.method == "POST":
                form = AddEmprunt(request.POST)
                if form.is_valid():
                    titre = form.cleaned_data["titre"]
                    loginMail = form.cleaned_data["loginMail"]
                    try:
                        doc = Documents.objects.get(titre = titre.split("-")[0].strip(), auteur = titre.split("-")[1].strip(), existe = True, disponible = True)
                    except:
                        tom, msg = ("info", "Impossible d'affecter un prêt à ce document")
                    else:
                        try:
                            user = Users.objects.get(loginMail = loginMail, state = True, elligible = True)
                        except:
                            tom, msg = ("info", "Impossible d'affecter un prêt à cet utilisateur")
                        else:
                            global random_word
                            idfy = "{}{}{}{}".format(loginMail, choice(range(100)), choice(random_word), titre)
                            idfy = idfy.encode()
                            idfy = sha1(idfy)
                            Emprunts(idfy = idfy.hexdigest(), user = user, book = doc, dateEmprunt = today).save()
                            user.elligible = False
                            user.save()
                            doc.disponible = False
                            doc.save()
                            tom, msg = ("tip", "Le prêt a été affecter avec succès")
            else:
                form = AddEmprunt()
            return render(request, "admins/emp_index.html", locals())
        else:return redirect(f"/Accueil/{user_idf}")
    else:return redirect("/")

def renew(request, user_idf, emp_idfy):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        if isAdmin(user_idf):
            try:
                emp = Emprunts.objects.get(idfy = emp_idfy, regle = False)
            except Exception as e:
                print(e)
                return redirect(f"/Admin/{user_idf}/Emprunts/")
            else:
                today = date.today()
                emp.dateEmprunt = f"{today.day}/{today.month}/{today.year}"
                emp.save()
                searchGroup, user, doc = ("Emprunts", emp.user, emp.book)
                tom, msg = ("txt-s", "La date d'emprunt a été ramenée à celle d'aujourd'hui")
            return render(request, "search.html", locals())
        else:return redirect(f"/Accueil/{user_idf}")
    else:return redirect("/")

def fresh(request, user_idf, emp_idfy):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        if isAdmin(user_idf):
            try:
                emp = Emprunts.objects.get(idfy = emp_idfy, regle = False)
            except Exception as e:
                print(e)
                return redirect(f"/Admin/{user_idf}/Emprunts/")
            else:
                emp.regle = True
                emp.user.elligible = True
                emp.user.save()
                emp.book.disponible = True
                emp.book.save()
                emp.save()
                tom, msg = ("txt-s", "L'utilisateur a été affranchi de son emprunt")
            return render(request, "search.html", locals())
        else:return redirect(f"/Accueil/{user_idf}")
    else:return redirect("/")

def diagnostics(request, user_idf):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        if isAdmin(user_idf):
            typeOfDoc = []
            filiere = []
            emp = []
            emp_cat = []
            docs = Documents.objects.all()
            users = Users.objects.all()
            for i in docs:
                if not i.typeOfDoc in typeOfDoc:
                    typeOfDoc.append(i.typeOfDoc)
            for i in users:
                if not i.filiere in filiere:
                    if not i.filiere in ["-", ""]:
                        filiere.append(i.filiere)
            for i in Emprunts.objects.all():
                if not i in emp:
                    emp.append(i)
            for i in emp:
                if not i.book.typeOfDoc in emp_cat:
                    emp_cat.append(i.book.typeOfDoc)
            if len(docs.filter(existe = True)):
                plt.barh(["Existe", "Supprimé"], [len(docs.filter(existe = True)), len(docs.filter(existe = False))], color = ["green", "red"])
                plt.title("Graphe des documents en fonction de leur état")
                plt.savefig("cidop/index/templates/assets/img/gDocex.png")
                plt.close()
            
                plt.pie([len(docs.filter(existe = True)), len(docs.filter(existe = False))], labels=["Existe", "Supprimé"], autopct='%1.2f%%', shadow=True, startangle=90)
                plt.title("Graphe des documents en fonction de leur état")
                plt.savefig("cidop/index/templates/assets/img/gDocex_pie.png")
                plt.close()
                ##
                plt.bar([i[:3] + "..." for i in typeOfDoc], [len(docs.filter(typeOfDoc = i)) for i in typeOfDoc ])
                plt.title("Graphe des documents en fonction de leur catégorie")
                plt.savefig("cidop/index/templates/assets/img/gDocstate.png")
                plt.close()
                plt.pie([len(docs.filter(typeOfDoc = i)) for i in typeOfDoc ], labels = typeOfDoc, autopct='%1.2f%%', shadow=True, startangle=90)
                plt.title("Graphe des documents en fonction de leur catégorie")
                plt.savefig("cidop/index/templates/assets/img/gDocstate_pie.png")
                plt.close()
                ##
            plt.bar(["Réglé", "En cours"], [len(Emprunts.objects.filter(regle = True)), len(Emprunts.objects.filter(regle = False))], color=["green", "red"])
            plt.title("Graphe des emprunts en fonction de leur état")
            plt.savefig("cidop/index/templates/assets/img/gEmpex.png")
            plt.close()

            plt.pie([len(Emprunts.objects.filter(regle = True)), len(Emprunts.objects.filter(regle = False))], labels=["Réglé", "En cours"], autopct='%1.2f%%', shadow=True, startangle=90)
            plt.title("Graphe des emprunts en fonction de leur état")
            plt.savefig("cidop/index/templates/assets/img/gEmpex_pie.png")
            plt.close()
            
            plt.barh(["Existe", "Supprimé"], [len(users.filter(state = True)), len(users.filter(state = False))], color = ["green", "red"])
            plt.title("Graphe des utilisateurs en fonction de leur état")
            plt.savefig("cidop/index/templates/assets/img/gUserex.png")
            plt.close()
            plt.pie([len(users.filter(state = True)), len(users.filter(state = False))], labels=["Existe", "Supprimé"], autopct='%1.2f%%', startangle=90, shadow=True)
            plt.title("Graphe des utilisateurs en fonction de leur état")
            plt.savefig("cidop/index/templates/assets/img/gUserex_pie.png")
            plt.close()
            ##
            plt.barh(filiere, [len(users.filter(filiere = i)) for i in filiere])
            plt.title("Graphe des utilisateurs en fonction de leur filière")
            plt.savefig("cidop/index/templates/assets/img/gUserfil.png")
            plt.close()
            plt.pie([len(users.filter(filiere = i)) for i in filiere], labels=filiere, autopct='%1.2f%%', startangle=90)
            plt.title("Graphe des utilisateurs en fonction de leur filière")
            plt.savefig("cidop/index/templates/assets/img/gUserfil_pie.png")
            plt.close()
            ##


            return render(request, "admins/diagnostics.html", locals())
        else:return redirect(f"/Accueil/{user_idf}")
    else:return redirect("/")

def import_data(request, user_idf, import_group):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        if isAdmin(user_idf):
            if request.method == "POST":
                form = xlsFile(request.POST, request.FILES)
                if form.is_valid():
                    fle = form.cleaned_data["xls_file"]
                    xls(xls_file = fle).save()
                    xls_fle = xls.objects.order_by("-id")[0]
                    xls_fle = xls_fle.xls_file.path
                    if not xls_fle.split(".")[-1] in ["xls", "xlsx"]:
                        tom, msg = ("info", "Le format du fichier est invalide (.xls, .xlsx requis)")
                        return render(request, "admins/import_data.html", locals())
                    tom, msg = autoAdd(xls_fle, import_group)
            else:
                form = xlsFile()
            return render(request, "admins/import_data.html", locals())
        else:return redirect(f"/Accueil/{user_idf}")
    else:return redirect("/")

def doc_info(request, user_idf, doc_idfy):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        books_target = ""
        for i in Documents.objects.filter(existe = True):
            books_target += f"{i.titre} - {i.auteur}##"
        if request.GET.get("search") or not request.GET.get("search") in [None, ""]:
            doc = Documents.objects.get(titre = request.GET.get("search").split("-")[0].strip(), auteur = request.GET.get("search").split("-")[1].strip())
            return redirect(f"/DocInfo/{user_idf}/{doc.idfy}/")
        user = userAllInfoFromUserIdf(user_idf)
        rsved = reserved(user_idf, doc_idfy)
        try:
            doc = Documents.objects.get(idfy = doc_idfy)
        except:
            doc = None
        else:pass
        return render(request, "doc_info.html", locals())
    else:return redirect("/")

def fake_page(request, user_idf, target):
    if target.__contains__("reserve+"):
        doc_idfy = target.split("+")[1].strip()
        doc = Documents.objects.get(idfy=doc_idfy)
        user = userAllInfoFromUserIdf(user_idf)
        if not user.idfy in doc.reservedby:
            doc.reservedby += user.idfy + ";"
            doc.save()
    if target.__contains__("unresv+"):
        doc_idfy = target.split("+")[1].strip()
        doc = Documents.objects.get(idfy=doc_idfy)
        user = userAllInfoFromUserIdf(user_idf)
        if user.idfy in doc.reservedby:
            doc.reservedby = doc.reservedby.replace(user.idfy + ";", "")
            doc.save()
    if target.__contains__("user_delete+"):
        user_idfy = target.split("+")[1].strip()
        try:
            user = Users.objects.get(idfy=user_idfy, state = True)
        except Exception as e:
            print(e)
        else:
            user.state = False
            user.save()
    if target.__contains__("doc_delete+"):
        doc_idfy = target.split("+")[1].strip()
        try:
            doc = Documents.objects.get(idfy=doc_idfy, existe = True)
        except Exception as e:
            print(e)
        else:
            doc.existe = False
            doc.save()

    if target.__contains__("doc_restaure+"):
        doc_idfy = target.split("+")[1].strip()
        try:
            doc = Documents.objects.get(idfy=doc_idfy, existe = False)
        except Exception as e:
            print(e)
        else:
            doc.existe = True
            doc.save()

    if target.__contains__("user_restaure+"):
        user_idfy = target.split("+")[1].strip()
        try:
            user = Users.objects.get(idfy=user_idfy, state = False)
        except Exception as e:
            print(e)
        else:
            user.state = True
            user.save()

    if target.__contains__("renew+"):
        emp_idfy = target.split("+")[1].strip()
        try:
            emp = Emprunts.objects.get(idfy = emp_idfy, regle = False)
        except Exception as e:
            print(e)
        else:
            emp.dateEmprunt = date.today()
            emp.save()

    if target.__contains__("aff_emp+"):
        emp_idfy = target.split("+")[1].strip()
        try:
            emp = Emprunts.objects.get(idfy = emp_idfy, regle = False)
        except Exception as e:
            print(e)
        else:
            emp.regle = True
            emp.user.elligible = True
            emp.user.save()
            emp.book.disponible = True
            emp.book.save()
            emp.save()

    if target.__contains__("rest_emp+"):
        emp_idfy = target.split("+")[1].strip()
        try:
            emp = Emprunts.objects.get(idfy = emp_idfy, regle = True)
        except Exception as e:
            print(e)
        else:
            emp.regle = False
            emp.user.elligible = True
            emp.user.save()
            emp.book.disponible = True
            emp.book.save()
            emp.save()
    return render(request, "fake_page.html", locals())

def export_data(request, user_idf, export_group):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        if isAdmin(user_idf):
            if export_group == "Utilisateurs":
                form = exportUser()
                opt = request.GET.get("opt")
                if opt and not opt in ["", None]:
                    if makeExportXLS(export_group, opt):
                        return redirect(f"/static/xls_file/export_{export_group}_{opt}.xlsx")
                    else:
                        msg = "Une erreur est survenue lors de l'exportation des données"
            elif export_group == "Documents":
                form = exportDoc()
                opt = request.GET.get("opt")
                if opt and not opt in ["", None]:
                    if makeExportXLS(export_group, opt):
                        return redirect(f"/static/xls_file/export_{export_group}_{opt}.xlsx")
                    else:
                        msg = "Une erreur est survenue lors de l'exportation des données"
            return render(request, "admins/export_data.html", locals())
        else:
            return redirect(f"/Accueil/{user_idf}")
    else:
        return redirect("/")


def acte_emp(request, user_idf, emp_idfy):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        try:
            emp = Emprunts.objects.get(idfy = emp_idfy)
        except:
            return redirect(f"/Admin/{user_idf}/Emprunts/")
        deadLine = add2Weeks(emp.dateEmprunt)
        return render(request, "admins/acte_demp.html", locals())
    else:return redirect("/")

def facture(request, user_idf, emp_idfy):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        try:
            emp = Emprunts.objects.get(idfy = emp_idfy)
        except:
            return redirect(f"/Admin/{user_idf}/Emprunts/")
        deadLine = add2Weeks(emp.dateEmprunt)
        penDay = retard(emp.dateEmprunt, deadLine)
        penVal = calculatePenality(penDay)
        return render(request, "admins/facture.html", locals())
    else:return redirect("/")

def read_doc(request, user_idf, doc_idfy):
    projet_name = 'CIDOP'
    global areConnected
    if isConnected(user_idf):
        adm = False
        if isAdmin(user_idf):
            adm = True
        try:
            doc = Documents.objects.get(idfy = doc_idfy)
        except:
            return redirect(request.GET.get("old"))
        passwd = doc_idfy[3:16]
        old = request.GET.get("old")
        canApp = True if len(Appreciations.objects.filter(user = userAllInfoFromUserIdf(user_idf), book = doc)) == 0 else False
        if request.method == "POST":
            form = Appreciation(request.POST)
            if form.is_valid():
                note = form.cleaned_data["note"]
                app = Appreciations()
                app.book, app.user, app.note = doc, userAllInfoFromUserIdf(user_idf), note
                app.save()
                apps = Appreciations.objects.filter(book = doc)
                notes = []
                for i in apps:
                    notes.append(i.note)
                doc.avg = sum(notes) / len(notes)
                doc.save()
        else:
            form = Appreciation()
        return render(request, "read_doc.html", locals())
    else:return redirect("/")

## Les views relatives à l'API de cidop

def api_login(request):
    loginMail = request.GET.get("loginMail")
    mdp = request.GET.get("mdp")
    if loginMail and not loginMail in ["", None]:
        if mdp and not mdp in ["", None]:
            try:
                user = Users.objects.get(loginMail = loginMail, mdp = mdp, state = True)
            except Exception as e:
                return JsonResponse({"error": "L'utilisateur est introuvable"})
            else:
                global areConnceted
                if not user in areConnceted:
                    areConnceted.append(user)
                return JsonResponse({
                    "user_idf": makeUserIdf(user.idfy),
                    "idfy": user.idfy, 
                    "nom": user.nom,
                    "prenom": user.prenom,
                    "addrMail": user.addrMail,
                    "loginMail": user.loginMail,
                    "mdp": user.mdp,
                    "filiere": user.filiere,
                    "elligible": "true" if user.elligible else "false",
                    })
        return JsonResponse({"error": "Veuillez renseigner les informations"})
    return JsonResponse({"error": "Veuillez renseigner les informations"})

def api_all_docs(request, user_idf):
    if isConnected(user_idf):
        docs = Documents.objects.filter(existe = True).order_by("titre")
        data = []
        for i in docs:
            data.append({
                "idfy": i.idfy,
                "titre": i.titre,
                "resume": i.resume,
                "typeOfDoc": i.typeOfDoc.capitalize(),
                "auteur": i.auteur,
                "reservedBy": i.reservedby,
                "isbn": i.isbn,
                "physique": "true" if i.physique else "false",
                "emplacement": i.file_path.url if not i.physique else i.emplacement,
                "disponible": "true" if i.disponible else "false",
                "avg": i.avg
                })
        return JsonResponse({"data": data})
    else:
        return JsonResponse({"error": "Veuillez vous connecter au préalable"})

def api_recents_docs(request, user_idf):
    if isConnected(user_idf):
        data = []
        for i in Documents.objects.order_by("-id")[:15]:
            data.append({
                "idfy": i.idfy,
                "titre": i.titre,
                "resume": i.resume,
                "typeOfDoc": i.typeOfDoc.capitalize(),
                "auteur": i.auteur,
                "reservedBy": i.reservedby,
                "isbn": i.isbn,
                "physique": "true" if i.physique else "false",
                "emplacement": i.file_path.url if not i.physique else i.emplacement,
                "disponible": "true" if i.disponible else "false",
                "avg": i.avg
                })
        return JsonResponse({"data": data})
    else:
        return JsonResponse({"error": "Veuillez vous connecter au préalable"})

def api_most_noted(request, user_idf):
    if isConnected(user_idf):
        docs, data = [], []
        for i in Documents.objects.filter(physique = False, existe = True):
            if i .avg >= 12:
                if not i in docs:
                    docs.append(i)
        for i in docs:
            data.append({
                "idfy": i.idfy,
                "titre": i.titre,
                "resume": i.resume,
                "typeOfDoc": i.typeOfDoc.capitalize(),
                "auteur": i.auteur,
                "reservedBy": i.reservedby,
                "isbn": i.isbn,
                "physique": "true" if i.physique else "false",
                "emplacement": i.file_path.url if not i.physique else i.emplacement,
                "disponible": "true" if i.disponible else "false",
                "avg": i.avg
                })
        return JsonResponse({"data": data})
    else:
        return JsonResponse({"data": "Veuillez vous connecter au préalable"})

def api_recom(request, user_idf):
    if isConnected(user_idf):
        mine, docs, data = [], [], []
        if len(Appreciations.objects.all()) > 0:
            for i in Appreciations.objects.filter(user = userAllInfoFromUserIdf(user_idf)):
                    if i.note >= 12:
                        if not i.book.typeOfDoc in mine:
                            mine.append(i.book.typeOfDoc)
            for i in mine:
                for x in Documents.objects.filter(typeOfDoc = i):
                    docs.append(x)
            for i in docs:
                data.append({
                "idfy": i.idfy,
                "titre": i.titre,
                "resume": i.resume,
                "typeOfDoc": i.typeOfDoc.capitalize(),
                "auteur": i.auteur,
                "reservedBy": i.reservedby,
                "isbn": i.isbn,
                "physique": "true" if i.physique else "false",
                "emplacement": i.file_path.url if not i.physique else i.emplacement,
                "disponible": "true" if i.disponible else "false",
                "avg": i.avg
                })
        return JsonResponse({"data": data})
    else:
        return JsonResponse({"error": "Veuillez vous connecter au préalable"})

def api_mdpForgot(request):
    loginMail = request.GET.get("loginMail")
    recover = request.GET.get("recover")
    option = request.GET.get("recover")
    try:
        user = Users.objects.get(loginMail = loginMail, state = True)
    except:
        return JsonResponse({"error": "L'utilisateur est introuvable"})
    else:
        if user.idfy[3:11] == recover:
            if option or not option in ["", None]:
                if option == "disc_all":
                    global areConnceted
                    if user in areConnceted:
                        areConnceted.remove(user)
                user.mdp = recover
                user.save()
                return JsonResponse({"info": "Le mot de passe a été réinitilisé avec succès"})
            else:
                return JsonResponse({"error": "Vous devez choisir une option"})
        else:
            return JsonResponse({"error": "La clef de récupération est incorrecte"})

def api_firstConn(request, user_idf):
    if isConnected(user_idf):
        user = userAllInfoFromUserIdf(user_idf)
        if user == None:return JsonResponse({"error": "L'utilisateur est introuvable"})
        else:
            if user.mdp == user.idfy[3:11]:
                        mdp1 = request.GET.get('mdp1')
                        mdp2 = request.GET.get('mdp2')
                        print(mdp1)
                        print(mdp2)
                        if mdp1 != mdp2:return JsonResponse({"error": "Les mots de passes ne concordent pas"})
                        else:
                            if len(mdp1) < 8:return JsonResponse({"error" :"Mots de passes trop court."})
                            else:
                                if re.match(".{0,}[A-Za-z]+.{0,}", mdp1) == None:return JsonResponse({"error": "Doit contenir au moins une lettre"})
                                else:
                                    if re.match(".+[0-9]+.{0,}", mdp1) == None:return JsonResponse({"error": "Doit contenir au moins un chiffre"})
                                    else:
                                        if re.match(".{0,}[\@\_\+\.\-\#]+.{0,}", mdp1) == None:return JsonResponse({"error": "Doit contenir au moins '@', '_', '+', '.', '#', ou '-'"})
                                        else:
                                            user.mdp = mdp1
                                            user.save()
                                            return JsonResponse({"go": "true"})
            else:return JsonResponse({"go": "true"})
    else:print("Not connected")

def api_mdpForgot(request):
    projet_name = "CIDOP"
    loginMail = request.GET.get("loginMail")
    recover = request.GET.get("recover")
    try:user = Users.objects.get(loginMail = loginMail.strip())
    except:return JsonResponse({"error": "Une erreur s'est produite"})
    else:
        if user.idfy[3:11] == recover:    
            user.mdp = user.idfy[3:11]
            user.save()
            global areConnceted
            try:
                areConnceted.remove(user)
            except:
                pass
            return JsonResponse({"data":"Le mot de passe a été réinitialisé avec succès. Veuillez vous connecter avec votre clé secrète"})
        else:
            return JsonResponse({"error": "Une erreur s'est produite"})
    return render(request, 'mdpForgot.html', locals())

def api_appreciate_doc(request, user_idf):
    if isConnected(user_idf):
        user = userAllInfoFromUserIdf(user_idf)
        if user == None:return JsonResponse({"error": "L'utilisateur est introuvable"})
        else:
            doc_idfy = request.GET.get("doc_idfy")
            note = request.GET.get("note")
            try:doc = Documents.objects.get(idfy = doc_idfy)
            except:return JsonResponse({"error": "Le document est introuvable"})
            else:
                try:note = int(note)
                except:return JsonResponse({"error": "Assurez vous d'entrer un nombre"})
                else:
                    if note > 20:return JsonResponse({"error": "Vous ne devez pas déppaser 20"})
                    elif note < 0:return JsonResponse({"error": "Vous ne devez pas descendre en dessous de 0"})
                    else:
                        canApp = True if len(Appreciations.objects.filter(user = userAllInfoFromUserIdf(user_idf), book = doc)) == 0 else False
                        if canApp:
                            app = Appreciations()
                            app.book, app.user, app.note = doc, userAllInfoFromUserIdf(user_idf), note
                            app.save()
                            apps = Appreciations.objects.filter(book = doc)
                            notes = []
                            for i in apps:
                                notes.append(i.note)
                            doc.avg = sum(notes) / len(notes)
                            doc.save()
                            return JsonResponse({"data": "La note a été attribuée avec succès"})
                        else:
                            return JsonResponse({"error": "Vous avez déjà attribuer une note à ce document"})
